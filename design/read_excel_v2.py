import sys
sys.stdout.reconfigure(encoding='utf-8')
from openpyxl import load_workbook

def read_plans(filepath, pet_type):
    wb = load_workbook(filepath, data_only=True)
    ws = wb.active
    plans = []
    # 从第3行开始（跳过标题和第2行表头），最多10个方案
    row_idx = 3
    while row_idx <= 30:
        row = list(ws.iter_rows(min_row=row_idx, max_row=row_idx, values_only=True))[0]
        if not row or not row[0]:
            break
        pid = int(row[0]) if str(row[0]).isdigit() else row[0]
        name = row[1] or ''
        conditions = (row[2] or '').strip()
        ingredients_raw = (row[3] or '').strip()
        ingredient_weights = (row[4] or '').strip()
        daily_grams = (row[6] or '').strip()
        monthly_price = (row[7] or '').strip()
        effect = (row[8] or '').strip()
        remark = (row[9] or '').strip()

        # 解析8种成分名称
        ingredients = []
        if ingredients_raw:
            for line in str(ingredients_raw).split('\n'):
                line = line.strip()
                if not line:
                    continue
                # 去掉行首序号 "1. " 或 "1、"
                import re
                line = re.sub(r'^\d+[、.．]\s*', '', line)
                ingredients.append(line)

        # 解析对应克数/颗数
        weights = []
        if ingredient_weights:
            for line in str(ingredient_weights).split('\n'):
                line = line.strip()
                if not line:
                    continue
                line = re.sub(r'^\d+[、.．]\s*', '', line)
                weights.append(line)

        # 构建 ingredients 对象数组
        ing_objs = []
        for i, ing in enumerate(ingredients):
            w = weights[i] if i < len(weights) else ''
            ing_objs.append({'name': ing, 'weight': w})

        # 解析适用条件，构建 matchRules
        match_rules = {'isDefault': False}
        
        # 年龄
        if '幼' in conditions and '幼犬' in conditions or '0～12月' in conditions or '幼猫' in conditions:
            if '幼犬0～12月' in conditions or '幼犬' in conditions:
                match_rules['age'] = ['幼犬', '幼']
            if '幼猫' in conditions or '0～12月' in conditions:
                match_rules['age'] = ['幼猫', '幼']
        elif '老年' in conditions or '8岁' in conditions or '7岁' in conditions or '8岁+' in conditions or '7岁+' in conditions:
            if '老年犬' in conditions or '8岁' in conditions:
                match_rules['age'] = ['老年犬', '老']
            if '老年猫' in conditions or '7岁' in conditions:
                match_rules['age'] = ['老年猫', '老']
        else:
            match_rules['age'] = None  # 不限制=通用

        # 绝育
        if '已绝育' in conditions and '未绝育' not in conditions:
            match_rules['neuter'] = ['已绝育']
        elif '未绝育' in conditions and '已绝育' not in conditions:
            match_rules['neuter'] = ['未绝育']
        else:
            match_rules['neuter'] = None

        # 活动量
        if '高' in conditions and '活动量' in conditions:
            match_rules['activity'] = ['高']
        elif '低' in conditions and '活动量' in conditions:
            match_rules['activity'] = ['低']
        else:
            match_rules['activity'] = None

        # 饲养模式
        if '多宠' in conditions and '单宠' not in conditions:
            match_rules['mode'] = ['多宠']
        elif '单宠' in conditions and '多宠' not in conditions:
            match_rules['mode'] = ['单宠']
        else:
            match_rules['mode'] = None

        # 遛弯（狗）
        if 'walk' in filepath or pet_type == 'dog':
            if '＜1次' in conditions or '1～2次' in conditions:
                if '＜1次' in conditions:
                    match_rules['walk'] = ['＜1次']
                else:
                    match_rules['walk'] = ['1～2次']
            elif '3～5次' in conditions or '每天' in conditions:
                match_rules['walk'] = ['3～5次', '每天1次以上']
            else:
                match_rules['walk'] = None
        else:
            match_rules['walk'] = None

        # 健康条件 - 从conditions中提取
        # 肠胃
        if '肠胃健康' in conditions and ('软便' not in conditions) and ('便秘' not in conditions) and ('拉稀' not in conditions) and ('挑食' not in conditions) and ('积食' not in conditions):
            match_rules['stomach'] = None  # 健康，不特殊匹配
        elif '软便' in conditions or '肠胃敏感' in conditions or '拉稀' in conditions or '挑食' in conditions or '积食' in conditions:
            match_rules['stomach'] = ['软便', '拉稀', '挑食', '积食']
        else:
            match_rules['stomach'] = None

        # 皮肤
        if '皮肤健康' in conditions and ('掉毛' not in conditions) and ('皮屑' not in conditions) and ('瘙痒' not in conditions):
            match_rules['skin'] = None
        elif '掉毛' in conditions or '皮屑' in conditions or '瘙痒' in conditions or '异位性' in conditions or '脱毛' in conditions:
            match_rules['skin'] = ['掉毛重', '皮屑瘙痒', '毛发干枯', '异位性皮炎', '季节性脱毛', '瘙痒泛红']
        else:
            match_rules['skin'] = None

        # 泌尿（猫）
        if pet_type == 'cat':
            if '泌尿健康' in conditions and ('尿频' not in conditions) and ('泪痕' not in conditions):
                match_rules['urinary'] = None
            elif '尿频' in conditions or '泪痕' in conditions or '上火' in conditions:
                match_rules['urinary'] = ['尿频上火', '泪痕重']
            else:
                match_rules['urinary'] = None
        else:
            match_rules['urinary'] = None

        # 关节
        if '关节健康' in conditions and '关节老化' not in conditions and '关节' not in conditions.split('健康')[1] if '健康' in conditions else True:
            match_rules['joint'] = None
        elif '关节老化' in conditions or '关节' in conditions:
            match_rules['joint'] = ['关节老化']
        else:
            match_rules['joint'] = None

        # 过敏
        if '无过敏' in conditions and '过敏' not in conditions.replace('无过敏', ''):
            match_rules['allergy'] = None
        elif '过敏' in conditions:
            match_rules['allergy'] = ['肉类过敏', '谷物过敏']
        else:
            match_rules['allergy'] = None

        # 代谢（狗）
        if pet_type == 'dog':
            if '代谢正常' in conditions and '肥胖' not in conditions and '高血脂' not in conditions:
                match_rules['metabolism'] = None
            elif '肥胖' in conditions or '高血脂' in conditions:
                match_rules['metabolism'] = ['肥胖高血脂']
            else:
                match_rules['metabolism'] = None
        else:
            match_rules['metabolism'] = None

        # 泪痕（狗）
        if pet_type == 'dog':
            if '泪痕' in conditions or '上火' in conditions:
                match_rules['tear'] = ['顽固泪痕', '易上火']
            else:
                match_rules['tear'] = None

        # 判断是否为默认方案
        is_default = ('通用' in name and '全年龄段' in name) or ('全年龄段' in name and '基础' in name)
        if is_default:
            match_rules['isDefault'] = True
            # 默认方案不设置具体匹配规则
            for k in list(match_rules.keys()):
                if k != 'isDefault':
                    match_rules[k] = None

        plans.append({
            'id': pid,
            'name': name,
            'matchRules': match_rules,
            'ingredients': ing_objs,
            'dailyGrams': daily_grams.replace('g', '').replace('G', '').strip() if daily_grams else '',
            'monthlyPrice': int(monthly_price.replace('¥', '').replace('/月', '').strip()) if monthly_price and monthly_price.replace('¥', '').replace('/月', '').strip().isdigit() else 89,
            'effect': effect,
            'remark': remark,
        })
        row_idx += 1

    return plans

dog_plans = read_plans('static/梵优茗宠狗狗配餐10款搭配方案_8种成分版_豆包AI生成.xlsx', 'dog')
cat_plans = read_plans('static/梵优茗宠猫咪配餐10款搭配组合方案_豆包AI生成.xlsx', 'cat')

# 输出为JS
def rules_js(rules):
    parts = []
    for k, v in rules.items():
        if v is None:
            parts.append(f'        {k}: null')
        elif isinstance(v, bool):
            parts.append(f'        {k}: {str(v).lower()}')
        elif isinstance(v, list):
            items = ', '.join([f"'{x}'" for x in v])
            parts.append(f'        {k}: [{items}]')
        else:
            parts.append(f'        {k}: {v}')
    return ',\n'.join(parts)

def ings_js(ings):
    parts = []
    for ing in ings:
        parts.append(f"        {{ name: '{ing[\"name\"]}', weight: '{ing[\"weight\"]}' }}")
    return ',\n'.join(parts)

lines = []
lines.append('/**')
lines.append(' * 梵优茗宠 - 配餐方案数据（由Excel自动生成）')
lines.append(' */')
lines.append('')
lines.append('const MEAL_PLANS = {')

lines.append('  dog: [')
for p in dog_plans:
    lines.append('    {')
    lines.append(f"      id: {p['id']},")
    lines.append(f"      name: '{p['name']}',")
    lines.append('      matchRules: {')
    lines.append(rules_js(p['matchRules']))
    lines.append('      },')
    lines.append('      ingredients: [')
    lines.append(ings_js(p['ingredients']))
    lines.append('      ],')
    lines.append(f"      dailyGrams: '{p['dailyGrams']}',")
    lines.append(f"      monthlyPrice: {p['monthlyPrice']},")
    lines.append(f"      effect: '{p['effect']}',")
    lines.append(f"      remark: '{p['remark']}',")
    lines.append('    },')
lines.append('  ],')
lines.append('')

lines.append('  cat: [')
for p in cat_plans:
    lines.append('    {')
    lines.append(f"      id: {p['id']},")
    lines.append(f"      name: '{p['name']}',")
    lines.append('      matchRules: {')
    lines.append(rules_js(p['matchRules']))
    lines.append('      },')
    lines.append('      ingredients: [')
    lines.append(ings_js(p['ingredients']))
    lines.append('      ],')
    lines.append(f"      dailyGrams: '{p['dailyGrams']}',")
    lines.append(f"      monthlyPrice: {p['monthlyPrice']},")
    lines.append(f"      effect: '{p['effect']}',")
    lines.append(f"      remark: '{p['remark']}',")
    lines.append('    },')
lines.append('  ]')
lines.append('};')

print('\n'.join(lines))
